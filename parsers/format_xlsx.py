#!/usr/bin/env python3
"""
format_xlsx.py — Aplica formato visual a datosfinales.xlsx para facilitar
la edición y navegación manual. No modifica los datos.

Uso:
    python3 parsers/format_xlsx.py
    npm run format:xlsx
"""

import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.views import Selection

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
ROOT_DIR   = os.path.dirname(SCRIPT_DIR)
XLSX       = os.path.join(ROOT_DIR, "datosfinales.xlsx")

CATEGORY_COLORS = {
    "AGE":           "FFE3F0FB",
    "Autonomías":    "FFE8F5E9",
    "Gobierno":      "FFFCE4EC",
    "Congreso":      "FFFFF8E1",
    "Senado":        "FFF3E5F5",
    "Partidos":      "FFFFE0B2",
    "Universidades": "FFE0F7FA",
}

COLUMN_WIDTHS = [15, 40, 22, 15, 28, 15, 48, 15, 35, 25, 15, 28, 20, 20, 20]

HEADER_NAMES = [
    "Categoría", "Nombre", "Twitter", "Twitter Activo",
    "Bluesky", "Bluesky Activo", "Mastodon", "Mastodon Activo",
    "Email", "Detalle", "Tipo", "Grupo", "Circunscripción", "CCAA", "Partido",
]

HEADER_FILL  = PatternFill("solid", fgColor="FF1E293B")
HEADER_FONT  = Font(bold=True, color="FFFFFFFF", name="Calibri", size=11)
HEADER_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=False)

THIN_SIDE   = Side(style="thin", color="FFB0BEC5")
CELL_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)

DATA_FONT  = Font(name="Calibri", size=10)
DATA_ALIGN = Alignment(vertical="center", wrap_text=False)


def main():
    print(f"Cargando {XLSX}…")
    wb = openpyxl.load_workbook(XLSX)
    ws = wb["Sheet1"]

    ws.row_dimensions[1].height = 24
    for col_idx, name in enumerate(HEADER_NAMES, start=1):
        cell = ws.cell(1, col_idx)
        cell.value     = name
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = HEADER_ALIGN
        cell.border    = CELL_BORDER

    for col_idx, width in enumerate(COLUMN_WIDTHS, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    for row in range(2, ws.max_row + 1):
        cat   = ws.cell(row, 1).value
        color = CATEGORY_COLORS.get(cat, "FFFFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        ws.row_dimensions[row].height = 18
        for col_idx in range(1, 16):
            cell           = ws.cell(row, col_idx)
            cell.fill      = fill
            cell.font      = DATA_FONT
            cell.alignment = DATA_ALIGN
            cell.border    = CELL_BORDER

    ws.freeze_panes           = "A2"
    ws.auto_filter.ref        = f"A1:O{ws.max_row}"
    ws.sheet_view.zoomScale   = 90
    ws.sheet_view.topLeftCell = "A1"
    ws.sheet_view.selection   = [
        Selection(pane="topLeft",    activeCell="A1", sqref="A1"),
        Selection(pane="bottomLeft", activeCell="A2", sqref="A2"),
    ]

    wb.save(XLSX)
    print("Hecho.")
    print("  · Cabecera fija (freeze en A2)")
    print("  · Autofiltro en todas las columnas")
    print("  · Colores por categoría")
    print("  · Anchos de columna ajustados")


if __name__ == "__main__":
    main()
